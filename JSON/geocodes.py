'''
Construct geocode dictionary object by reading excel file and output to JSON
'''

import openpyxl
import json

wb = openpyxl.load_workbook('JSON/geo.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

dictionary = {}
i = 2
for row in sheet['A2':'A754']:
    for cellObj in row:
        print(cellObj.value)
        if dictionary.get(cellObj.value):
            pass 
            dictionary[cellObj.value].append(sheet.cell(row = i, column = 2).value)
        else:
            dictionary[cellObj.value] = [sheet.cell(row = i, column = 2).value]
        i += 1

json_val = json.dumps(dictionary)

with open('JSON/test.json','w') as file:
    #file.truncate()
    file.writelines(str(json_val))