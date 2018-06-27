import openpyxl

wb = openpyxl.load_workbook('excel/Employee.xlsx')
print(wb.get_sheet_names())
#sheet = wb.sheetnames()
sheet = wb.get_sheet_by_name('Personal Details')

wb.create_sheet('New Sheet')
sheet1 = wb.get_sheet_by_name('New Sheet')

tpl = ()
tpl = tuple(sheet['A1':'E5'])
print(tpl)

print('*** Printing Values of each Cell ***')
for rowOfCellObjects in tpl:
    values = []
    i = 0
    for cellObj in rowOfCellObjects:
        # row, col = i, 1
        #sheet.cell(row,col)
        #sheet1['A'+str(j)]=  cellObj.value
        values.insert(i,cellObj.value)
        i+=1
    print (values[:5])
    sheet1.append(values)

#print(sheet['A1'].value)
#sheet2['A1'] = sheet['A1'].value

wb.save('excel/emp.xlsx')
wb.close()



#c= sheet['B1']
#print(c.coordinate + ' ' + c.value)

