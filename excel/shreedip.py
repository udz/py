import openpyxl

wb = openpyxl.load_workbook('excel/statement1.xlsx')
sheet = wb.get_sheet_by_name('AccountStatement')
sheet1 = wb.get_sheet_by_name('Sheet1')
sheet2 = wb.create_sheet('Result')

print('\n')
values2 = []
for rowOfCellObjects in sheet1['A2':'A248']:
    i = 0
    for cellObj in rowOfCellObjects:
        values2.insert(i,cellObj.value)
        i+=1

'''
# CSV file output version
file = open('excel/output.csv','w')

for i in range (2209):
    #cellValue = sheet['C'+str(i)]
    dateTxn = sheet.cell(int(i+1),1).value
    dateVal = sheet.cell(int(i+1),2).value
    description = sheet.cell(int(i+1),3).value
    deposit = sheet.cell(int(i+1),7).value

    for value in values2:
        if description.find(str(value)) != -1:
            print(str(value) + " is in " + description + " Deposit Amt " + str(deposit) + " Transaction Date: " + dateTxn + " Value Date: " + dateVal)
            line = str(value) + "," + description + "," + str(deposit) + "," + dateTxn + "," + dateVal
            file.writelines(line)
            file.write("\n")
file.close()
'''

# Excel File Version
header = ['Check','Date Txn','Date Value','Description','Check','Withdraw','Deposit','Balance']
sheet2.append(header[:8])

for i in range (2209):
    dateTxn = sheet.cell(int(i+1),1).value
    dateVal = sheet.cell(int(i+1),2).value
    description = sheet.cell(int(i+1),3).value
    check = sheet.cell(int(i+1),4).value
    withdraw = sheet.cell(int(i+1),5).value
    deposit = sheet.cell(int(i+1),6).value
    balance = sheet.cell(int(i+1),7).value

    j = 0    
    row = []
    for value in values2:
        if description.find(str(value)) != -1:
            row.insert(0,value)
            row.insert(1,dateTxn)
            row.insert(2,dateVal)
            row.insert(3,description)
            row.insert(4,check)
            row.insert(5,withdraw)
            row.insert(6,deposit)
            row.insert(7,balance)
            sheet2.append(row[:8])

wb.save('excel/new.xlsx')
wb.close()

'''
c= sheet['B1']
print(c.coordinate + ' ' + c.value)

print(tuple(sheet['A1':'C3']))

print('*** Printing Values of each Cell ***')
for rowOfCellObjects in sheet['A1':'C3']:
    values = []
    i = 0
    for cellObj in rowOfCellObjects:
        values.insert(i,cellObj.value)
        i+=1
    print (values[:3])
'''